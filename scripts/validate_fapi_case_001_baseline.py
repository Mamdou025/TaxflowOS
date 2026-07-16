from __future__ import annotations

import csv
import json
import math
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CASE = ROOT / "tests" / "fixtures" / "mapping" / "fapi" / "case-001"
CSV_PATH = CASE / "expected_mapping.csv"
MANIFEST_PATH = CASE / "case_manifest.json"
REPORT_PATH = CASE / "validation_report.json"

errors: list[str] = []
warnings: list[str] = []
checks: list[dict[str, Any]] = []


def record(name: str, passed: bool, detail: str) -> None:
    checks.append({"check": name, "passed": passed, "detail": detail})
    if not passed:
        errors.append(f"{name}: {detail}")


def equivalent(left: Any, right: Any) -> bool:
    if left in (None, "") and right in (None, ""):
        return True
    try:
        return math.isclose(float(left), float(right), rel_tol=1e-9, abs_tol=1e-6)
    except (TypeError, ValueError):
        return str(left) == str(right)


manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
    rows = list(csv.DictReader(handle))

record("row_count", len(rows) == manifest["atomic_expectations"], f"CSV={len(rows)}, manifest={manifest['atomic_expectations']}")
ids = [row["expectation_id"] for row in rows]
record("unique_expectation_ids", len(ids) == len(set(ids)), f"unique={len(set(ids))}, total={len(ids)}")
record("case_id", all(row["case_id"] == manifest["case_id"] for row in rows), "Every row must use the manifest case_id.")
record("accepted_status", all(row["accepted_status"] == "active" for row in rows), "Every baseline row must be active.")

controlled_treatments = set(manifest["controlled_treatments"])
actual_treatments = {row["expected_treatment"] for row in rows}
record("controlled_treatments", actual_treatments <= controlled_treatments, f"unexpected={sorted(actual_treatments - controlled_treatments)}")

controlled_reviews = set(manifest["controlled_review_requirements"])
actual_reviews = {row["review_requirement"] for row in rows}
record("controlled_review_requirements", actual_reviews <= controlled_reviews, f"unexpected={sorted(actual_reviews - controlled_reviews)}")

counts_by_affiliate = dict(Counter(row["affiliate_alias"] for row in rows))
record("affiliate_counts", counts_by_affiliate == manifest["counts_by_affiliate"], f"CSV={counts_by_affiliate}, manifest={manifest['counts_by_affiliate']}")
counts_by_treatment = dict(Counter(row["expected_treatment"] for row in rows))
record("treatment_counts", counts_by_treatment == manifest["counts_by_treatment"], f"CSV={counts_by_treatment}, manifest={manifest['counts_by_treatment']}")

rules_path = (CASE / manifest["files"]["rules"]).resolve()
record("rules_file_exists", rules_path.is_file(), str(rules_path))
record("reference_workpaper_exists", (CASE / manifest["files"]["reference_workpaper"]).is_file(), manifest["files"]["reference_workpaper"])
for correction_file in manifest["files"]["correction_rounds"]:
    record(f"correction_exists:{correction_file}", (CASE / correction_file).is_file(), correction_file)

workbooks_formula: dict[str, Any] = {}
workbooks_value: dict[str, Any] = {}
for row in rows:
    source_file = row["source_file"]
    source_path = CASE / source_file
    if not source_path.is_file():
        errors.append(f"Missing source file for {row['expectation_id']}: {source_file}")
        continue

    if row["source_role"] == "primary":
        expected_file = {
            "FA_1 SAS": "fa-01_trial_balance.xlsx",
            "FA_2 Corp": "fa-02_trial_balance.xlsx",
        }.get(row["affiliate_alias"])
        if source_file != expected_file:
            errors.append(f"Affiliate isolation violation for {row['expectation_id']}: {row['affiliate_alias']} -> {source_file}")

    source_required = row["source_required"].lower() == "true"
    if source_required and row["source_role"] != "primary":
        errors.append(f"Required source must be primary for {row['expectation_id']}")

    if not row["source_sheet"]:
        if source_required:
            errors.append(f"Missing source sheet for {row['expectation_id']}")
        continue

    if source_file not in workbooks_formula:
        workbooks_formula[source_file] = load_workbook(source_path, data_only=False, read_only=False)
        workbooks_value[source_file] = load_workbook(source_path, data_only=True, read_only=False)

    formula_book = workbooks_formula[source_file]
    value_book = workbooks_value[source_file]
    if row["source_sheet"] not in formula_book.sheetnames:
        errors.append(f"Missing sheet for {row['expectation_id']}: {source_file}!{row['source_sheet']}")
        continue

    sheet_formula = formula_book[row["source_sheet"]]
    sheet_value = value_book[row["source_sheet"]]
    label_cell = row["source_label_cell"]
    value_cell = row["source_value_cell"]

    if source_required and not label_cell:
        errors.append(f"Missing label cell for {row['expectation_id']}")
    elif label_cell and sheet_formula[label_cell].value in (None, ""):
        errors.append(f"Empty label cell for {row['expectation_id']}: {source_file}!{row['source_sheet']}!{label_cell}")

    if value_cell:
        formula_or_value = sheet_formula[value_cell].value
        cached_value = sheet_value[value_cell].value
        actual_value = cached_value if isinstance(formula_or_value, str) and formula_or_value.startswith("=") else formula_or_value
        if not equivalent(row["source_value"], actual_value):
            errors.append(f"Value mismatch for {row['expectation_id']}: CSV={row['source_value']!r}, workbook={actual_value!r}")
        expected_formula = formula_or_value if isinstance(formula_or_value, str) and formula_or_value.startswith("=") else ""
        if row["source_formula"] != expected_formula:
            errors.append(f"Formula mismatch for {row['expectation_id']}: CSV={row['source_formula']!r}, workbook={expected_formula!r}")

for row in rows:
    if row["must_cite_source"].lower() != "true":
        errors.append(f"Source citation disabled for {row['expectation_id']}")
    if row["expected_treatment"] == "map_to_engine" and not row["expected_engine_field"]:
        errors.append(f"Missing engine field for mapped row {row['expectation_id']}")
    if row["affiliate_alias"] == "FA_3 GmbH" and row["source_role"] != "benchmark":
        errors.append("FA_3 must remain benchmark-only because no primary source was supplied.")
    if row["source_role"] == "benchmark" and row["source_required"].lower() != "false":
        errors.append(f"Benchmark evidence cannot be required primary support for {row['expectation_id']}")

record("row_level_validation", not errors, "All rows resolve to allowed files, sheets, cells, values, formulas, treatments, and evidence roles." if not errors else f"{len(errors)} error(s) found.")

report = {
    "case_id": manifest["case_id"],
    "status": "pass" if not errors else "fail",
    "checks_run": len(checks),
    "expectations_validated": len(rows),
    "errors": errors,
    "warnings": warnings,
    "checks": checks,
}
REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
print(json.dumps(report, indent=2, ensure_ascii=False))
raise SystemExit(1 if errors else 0)
