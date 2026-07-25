from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CASE = ROOT / "tests" / "fixtures" / "mapping" / "fapi" / "case-001"
CSV_PATH = CASE / "expected_mapping.csv"
MANIFEST_PATH = CASE / "case_manifest.json"
README_PATH = CASE / "README.md"

PRIMARY_FILES = {
    "FA_1 SAS": "fa-01_trial_balance.xlsx",
    "FA_2 Corp": "fa-02_trial_balance.xlsx",
}

workbooks_formula = {
    name: load_workbook(CASE / filename, data_only=False, read_only=False)
    for name, filename in PRIMARY_FILES.items()
}
workbooks_value = {
    name: load_workbook(CASE / filename, data_only=True, read_only=False)
    for name, filename in PRIMARY_FILES.items()
}


def source_value(affiliate: str, sheet: str, cell: str) -> tuple[Any, Any]:
    if not sheet or not cell:
        return None, None
    formula = workbooks_formula[affiliate][sheet][cell].value
    value = workbooks_value[affiliate][sheet][cell].value
    if isinstance(formula, str) and formula.startswith("="):
        return value, formula
    return formula, None


def add_row(
    rows: list[dict[str, Any]],
    *,
    affiliate: str,
    row_type: str,
    source_role: str,
    source_file: str,
    source_sheet: str = "",
    label_cell: str = "",
    value_cell: str = "",
    label: str,
    treatment: str,
    engine_field: str = "",
    currency: str = "",
    review: str = "not_required",
    discrepancy: bool = False,
    source_required: bool = True,
    rationale_any: str = "",
    notes: str = "",
) -> None:
    expectation_id = f"FAPI-001-{len(rows) + 1:03d}"
    value: Any = None
    formula: Any = None
    if source_role == "primary" and source_sheet and value_cell:
        value, formula = source_value(affiliate, source_sheet, value_cell)

    rows.append(
        {
            "case_id": "fapi-case-001",
            "expectation_id": expectation_id,
            "affiliate_alias": affiliate,
            "expected_row_type": row_type,
            "source_role": source_role,
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_label_cell": label_cell,
            "source_value_cell": value_cell,
            "source_fact_label": label,
            "source_value": "" if value is None else value,
            "source_formula": "" if formula is None else formula,
            "currency": currency,
            "expected_treatment": treatment,
            "expected_engine_field": engine_field,
            "review_requirement": review,
            "discrepancy_expected": str(discrepancy).lower(),
            "source_required": str(source_required).lower(),
            "must_cite_source": "true",
            "accepted_status": "active",
            "rationale_keywords_any": rationale_any,
            "notes": notes,
            "lineage_reference": "correction_round_03.md",
        }
    )


rows: list[dict[str, Any]] = []

# FA_1 SAS: 19 reviewed rows normalized to 22 atomic expectations.
add_row(rows, affiliate="FA_1 SAS", row_type="identity", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A2", label="Legal/source name confirmation", treatment="context_only", currency="EUR", rationale_any="legal name|alias|entity", notes="Must confirm the source legal name and map it to FA_1 SAS before numeric mapping.")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A22", value_cell="B22", label="Total sales / total income", treatment="context_only", currency="EUR", rationale_any="reconciliation|denominator|context")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A11", value_cell="B11", label="42100 - Ventes", treatment="do_not_map", currency="EUR", rationale_any="active business|third-party|not FAPI")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A12", value_cell="B12", label="42110 - Prestations de services - Europe", treatment="do_not_map", currency="EUR", rationale_any="historical treatment|not FAPI|active business")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A13", value_cell="B13", label="42200 - Revenus - intercos", treatment="map_to_engine", engine_field="otherFapiIncome", currency="EUR", rationale_any="intercompany|historical treatment|FAPI")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A16", value_cell="B16", label="43100 - Variation TEC/RR", treatment="context_only", currency="EUR", rationale_any="WIP|unearned revenue|context")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A19", value_cell="B19", label="44220 - Autres revenus", treatment="manual_review_required", engine_field="otherFapiIncome", currency="EUR", review="question_required", rationale_any="nature unclear|review|other income")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A47", value_cell="B47", label="Total cost of sales", treatment="overhead_allocation_candidate", engine_field="generalExpenses", currency="EUR", review="candidate_review", rationale_any="allocation|FAPI and non-FAPI|overhead")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A39", value_cell="B39", label="52200 - Sous-traitants - intercos", treatment="deductible_expense_candidate", engine_field="generalExpenses", currency="EUR", review="candidate_review", rationale_any="directly connected|intercompany income|deductible")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A61", value_cell="B61", label="Total administrative salaries and social charges", treatment="overhead_allocation_candidate", engine_field="generalExpenses", currency="EUR", review="candidate_review", rationale_any="allocation|overhead")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A111", value_cell="B111", label="Total selling and administrative expenses", treatment="overhead_allocation_candidate", engine_field="generalExpenses", currency="EUR", review="candidate_review", rationale_any="allocation|overhead")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A43", value_cell="B43", label="53100 - Repas Projets", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", discrepancy=True, rationale_any="meals|repas|partial addback", notes="Atomic expansion of reviewed grouped meals row; bridge is incomplete until all related accounts are included.")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A44", value_cell="B44", label="53150 - Frais de voyage et déplacement Projets", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", discrepancy=True, rationale_any="travel meals|déplacement|partial addback", notes="Atomic expansion of reviewed grouped meals row.")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A73", value_cell="B73", label="Total - 64000 - MEALS", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", discrepancy=True, rationale_any="meals|repas|partial addback", notes="Missed in Eval 3; explicitly required by reviewer feedback.")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A88", value_cell="B88", label="68170 - Épicerie", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", discrepancy=True, rationale_any="épicerie|meals|partial addback", notes="Missed French-labelled account explicitly required by reviewer feedback.")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A117", value_cell="B117", label="82100 - Amortissement immob. Corporelles", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", rationale_any="depreciation|amortissement|addback")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A129", value_cell="B129", label="90010 - Impôts exigibles", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", rationale_any="accounting tax expense|addback|proration", notes="Rationale should mention that proration may be required if the expense relates to FAPI and non-FAPI activities.")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A121", value_cell="B121", label="83200 - Intérêts et pénalités", treatment="schedule_1_adjustment_candidate", engine_field="additions", currency="EUR", review="candidate_review", rationale_any="interest and penalties|intérêts|pénalités|addback")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A132", value_cell="B132", label="85500 - Échange devises réalisé", treatment="manual_review_required", engine_field="realized_fx_candidate", currency="EUR", review="question_required", rationale_any="realized FX|current account|capital account|income stream")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="ProfitandLoss", label_cell="A133", value_cell="B133", label="85660 - Unrealized Gain/Loss", treatment="do_not_map", currency="EUR", rationale_any="unrealized FX|do not map")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="BalanceSheet", label_cell="A32", value_cell="B32", label="15100 - Frais payés d'avance", treatment="manual_input_required", engine_field="additions or deductions", currency="EUR", review="question_required", rationale_any="prepaid|movement|prior year|current year")
add_row(rows, affiliate="FA_1 SAS", row_type="mapping", source_role="primary", source_file="fa-01_trial_balance.xlsx", source_sheet="BalanceSheet", label_cell="A90", value_cell="B90", label="27200 - Avance de Canco_1 Inc", treatment="client_question", currency="EUR", review="question_required", rationale_any="intercompany advance|interest|FX|terms")

# FA_2 Corp: 18 reviewed rows normalized to 21 atomic expectations.
add_row(rows, affiliate="FA_2 Corp", row_type="identity", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A2", label="Legal/source name confirmation", treatment="context_only", rationale_any="legal name|alias|entity", notes="Must confirm the source legal name and map it to FA_2 Corp before numeric mapping.")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A19", value_cell="B19", label="Total income", treatment="context_only", rationale_any="reconciliation|denominator|context")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A12", value_cell="B12", label="42100 - Ventes", treatment="do_not_map", rationale_any="active business|third-party|not FAPI")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A13", value_cell="B13", label="42200 - Revenus - intercos", treatment="map_to_engine", engine_field="otherFapiIncome", rationale_any="intercompany|historical treatment|FAPI")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A16", value_cell="B16", label="43100 - Variation TEC/RR", treatment="context_only", rationale_any="WIP|unearned revenue|context")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A35", value_cell="B35", label="Total cost of sales", treatment="overhead_allocation_candidate", engine_field="generalExpenses", review="candidate_review", rationale_any="allocation|FAPI and non-FAPI|overhead")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A43", value_cell="B43", label="Total administrative salaries and social charges", treatment="overhead_allocation_candidate", engine_field="generalExpenses", review="candidate_review", rationale_any="allocation|overhead")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A85", value_cell="B85", label="Total selling and administrative expenses", treatment="overhead_allocation_candidate", engine_field="generalExpenses", review="candidate_review", rationale_any="allocation|overhead")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A69", value_cell="B69", label="Facility expenses", treatment="overhead_allocation_candidate", engine_field="generalExpenses", review="candidate_review", rationale_any="facility|allocation|overhead")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A30", value_cell="B30", label="53100 - Repas Projets", treatment="schedule_1_adjustment_candidate", engine_field="additions", review="candidate_review", discrepancy=True, rationale_any="meals|repas|partial addback", notes="Atomic expansion of reviewed grouped meals row.")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A55", value_cell="B55", label="Total - 64000 - MEALS", treatment="schedule_1_adjustment_candidate", engine_field="additions", review="candidate_review", discrepancy=True, rationale_any="meals|repas|partial addback", notes="Atomic expansion of reviewed grouped meals row.")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A68", value_cell="B68", label="68170 - Épicerie", treatment="schedule_1_adjustment_candidate", engine_field="additions", review="candidate_review", discrepancy=True, rationale_any="épicerie|meals|partial addback", notes="Missed French-labelled account explicitly required by reviewer feedback.")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A92", value_cell="B92", label="82100 - Amortissement immob. Corporelles", treatment="schedule_1_adjustment_candidate", engine_field="additions", review="candidate_review", rationale_any="depreciation|amortissement|addback")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A93", value_cell="B93", label="82150 - Amortissement immob. Incorporelles", treatment="schedule_1_adjustment_candidate", engine_field="additions", review="candidate_review", rationale_any="amortization|amortissement|addback")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A102", value_cell="B102", label="90100 - Impôts U.S.", treatment="schedule_1_adjustment_candidate", engine_field="additions", review="candidate_review", rationale_any="accounting tax expense|addback|not foreign accrual tax")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A105", value_cell="B105", label="85500 - Échange devises réalisé", treatment="manual_review_required", engine_field="realized_fx_candidate", review="question_required", rationale_any="realized FX|current account|capital account|income stream")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="P&L", label_cell="A99", value_cell="B99", label="85550 - Rounding Gain/Loss", treatment="do_not_map", rationale_any="rounding|immaterial|do not map")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="BalanceSheet", label_cell="A25", value_cell="B25", label="15100 - Frais payés d'avance", treatment="manual_input_required", engine_field="additions or deductions", review="question_required", rationale_any="prepaid|movement|prior year|current year")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="benchmark", source_file="reference_fapi_workpaper.xlsx", source_sheet="FA_2 Corp - Info + Sch 1", label="Tax depreciation / CCA", treatment="manual_input_required", engine_field="tax_depreciation_manual_input", review="question_required", source_required=False, rationale_any="tax depreciation|CCA|detailed support|manual input", notes="Primary source support is absent; historical workpaper is benchmark evidence only.")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="BalanceSheet", label_cell="A19", value_cell="B19", label="12200 - Comptes clients - intercos", treatment="client_question", review="question_required", rationale_any="intercompany receivable|interest|FX|counterparty")
add_row(rows, affiliate="FA_2 Corp", row_type="mapping", source_role="primary", source_file="fa-02_trial_balance.xlsx", source_sheet="BalanceSheet", label_cell="A55", value_cell="B55", label="21020 - Comptes fournisseurs - intercos", treatment="client_question", review="question_required", rationale_any="intercompany payable|interest|FX|counterparty")

# FA_3 GmbH: coverage expectation because no standalone primary source was provided.
add_row(rows, affiliate="FA_3 GmbH", row_type="coverage", source_role="benchmark", source_file="reference_fapi_workpaper.xlsx", source_sheet="FA_3 GmbH 2024", label="Historical FAPI insights with missing primary source", treatment="benchmark_only", review="question_required", source_required=False, rationale_any="benchmark|source missing|historical treatment", notes="The agent must include FA_3, label historical insights as benchmark-only, and must not invent or borrow primary source facts.")

fieldnames = list(rows[0].keys())
with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
    writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)

counts_by_affiliate: dict[str, int] = {}
counts_by_treatment: dict[str, int] = {}
for row in rows:
    counts_by_affiliate[row["affiliate_alias"]] = counts_by_affiliate.get(row["affiliate_alias"], 0) + 1
    counts_by_treatment[row["expected_treatment"]] = counts_by_treatment.get(row["expected_treatment"], 0) + 1

manifest = {
    "case_id": "fapi-case-001",
    "workflow": "fapi",
    "status": "active_baseline",
    "baseline_source": "correction_round_03.md",
    "original_reviewed_rows": 38,
    "atomic_expectations": len(rows),
    "normalization_note": "Grouped meals and amortization rows were expanded into atomic source-level expectations. Missing identity and FA_3 coverage requirements are represented explicitly.",
    "reference_workpaper_role": "benchmark_only",
    "primary_source_role": "trial_balances",
    "target_period": 2024,
    "affiliates": [
        {
            "alias": "FA_1 SAS",
            "primary_file": "fa-01_trial_balance.xlsx",
            "source_name_contains": "Client A SAS",
            "currency": "EUR",
            "primary_source_available": True,
        },
        {
            "alias": "FA_2 Corp",
            "primary_file": "fa-02_trial_balance.xlsx",
            "source_name_contains": "Client A Corp",
            "currency": None,
            "primary_source_available": True,
            "currency_note": "Currency is not explicit in the source preview and should not be inferred.",
        },
        {
            "alias": "FA_3 GmbH",
            "primary_file": None,
            "source_name_contains": None,
            "currency": None,
            "primary_source_available": False,
            "benchmark_sheet": "FA_3 GmbH 2024",
        },
    ],
    "files": {
        "expected_mapping": "expected_mapping.csv",
        "rules": "../../../../../Tax rules/Fapi rules/FAPI_Final_Mapping_Rules.md",
        "reference_workpaper": "reference_fapi_workpaper.xlsx",
        "correction_rounds": [
            "correction_round_01.md",
            "correction_round_02.md",
            "correction_round_03.md",
        ],
    },
    "controlled_treatments": sorted(counts_by_treatment),
    "controlled_review_requirements": ["not_required", "candidate_review", "question_required"],
    "counts_by_affiliate": counts_by_affiliate,
    "counts_by_treatment": counts_by_treatment,
    "hard_fail_conditions": [
        "The agent uses historical workpaper summary values as primary engine inputs.",
        "The agent leaks facts between affiliates.",
        "The agent invents a primary source for FA_3 GmbH.",
        "The agent maps a target field not allowed by the active workflow schema.",
        "The agent omits source citations for a proposed mapping.",
    ],
    "scoring": {
        "required_keys": [
            "affiliate_alias",
            "source_fact_label",
            "expected_treatment",
            "expected_engine_field",
            "source_file",
            "source_sheet",
            "source_label_cell",
        ],
        "primary_match": "affiliate_alias + normalized source_fact_label or exact source location",
        "treatment_match": "exact",
        "engine_field_match": "exact when expected_engine_field is populated",
        "review_match": "exact controlled review requirement",
        "citation_match": "source file, sheet, and cell must resolve when source_required is true",
    },
}
MANIFEST_PATH.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

readme = f"""# FAPI Case 001 Evaluation Baseline

## Purpose

This folder is the first reproducible evaluation case for the FAPI source-mapping agent. The **trial balances are primary source evidence**. The historical FAPI workpaper is benchmark and logic-support evidence only and must not be used as a substitute for missing primary source values.

## Affiliate scope

| Affiliate | Primary source | Expected handling |
| --- | --- | --- |
| FA_1 SAS | `fa-01_trial_balance.xlsx` | Map and classify source facts in EUR, preserving exact sheet and cell citations. |
| FA_2 Corp | `fa-02_trial_balance.xlsx` | Map and classify source facts; do not infer currency when it is not explicit. |
| FA_3 GmbH | No standalone source supplied | Include benchmark-only historical insights and flag missing primary support. Do not invent or borrow facts. |

## Baseline files

| File | Role |
| --- | --- |
| `expected_mapping.csv` | Machine-readable expected output with one atomic expectation per row. |
| `case_manifest.json` | Case scope, controlled values, scoring rules, and hard-fail conditions. |
| `correction_round_01.md` to `correction_round_03.md` | Reviewer-feedback lineage. Round 3 is the accepted baseline. |
| `reference_fapi_workpaper.xlsx` | Historical benchmark and treatment evidence only. |
| Trial-balance workbooks | Primary source evidence. |

## Normalization

Eval 3 reviewed **38 rows**. The machine baseline contains **{len(rows)} atomic expectations** because grouped meals and amortization outputs were split into individual source facts. The baseline also represents missing legal-name confirmation rows and the FA_3 benchmark-only coverage requirement explicitly.

## Pass requirements

A future mapping run must identify the affiliate, source fact, treatment, target engine field when applicable, review requirement, and exact evidence citation. It must run post-mapping validation and must satisfy every hard-fail condition in `case_manifest.json`.

## Update policy

Do not overwrite accepted expectations silently. Changes should cite the reviewer feedback that supports them, update the manifest counts, and be committed as a new reviewed baseline version.
"""
README_PATH.write_text(readme, encoding="utf-8")

print(json.dumps({"expected_mapping": str(CSV_PATH), "manifest": str(MANIFEST_PATH), "readme": str(README_PATH), "rows": len(rows)}, indent=2))
